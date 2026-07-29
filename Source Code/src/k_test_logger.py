"""k_test_logger.py
==================
Logger ringan untuk tombol K di YOLO.py: tiap K ditekan, robot dikirim G1
ke titik hasil homography (lihat config.K_TEST_MOVE_Z_MM), lalu 1 baris
baru (Percobaan, Estimasi X, Estimasi Y) ditulis ke file XLSX di
config.K_TEST_LOG_XLSX_PATH.

Kolom Aktual X/Y sengaja dikosongkan -- diisi manual hasil ukur penggaris.
Kolom Error dihitung otomatis lewat rumus Excel (jarak Euclidean), jadi
otomatis terisi begitu Aktual X/Y diisi dan file dibuka/disimpan di Excel.

Modul ini berdiri sendiri (tidak menyentuh hasil_pick_place.csv/xlsx milik
esp32_comm.py) supaya tidak mengubah struktur file pengujian pick-place
yang sudah ada.
"""
from __future__ import annotations

import os
from typing import Any

HEADER_ROW1 = ["Percobaan", "Estimasi", "", "Aktual", "", "Error (mm)"]
HEADER_ROW2 = ["", "X", "Y", "X", "Y", ""]
FIRST_DATA_ROW = 3  # baris 1-2 = header 2 tingkat


def _log_path(config: Any) -> str:
    return str(getattr(config, "K_TEST_LOG_XLSX_PATH", os.path.join(os.getcwd(), "logs", "uji_akurasi_titik.xlsx")))


def _new_workbook(path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Akurasi Titik"

    font_name = "Times New Roman"
    header_font = Font(name=font_name, bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:A2")
    ws["A1"] = "Percobaan"
    ws.merge_cells("B1:C1")
    ws["B1"] = "Estimasi"
    ws.merge_cells("D1:E1")
    ws["D1"] = "Aktual"
    ws.merge_cells("F1:F2")
    ws["F1"] = "Error (mm)"
    ws["B2"] = "X"
    ws["C2"] = "Y"
    ws["D2"] = "X"
    ws["E2"] = "Y"

    for row in (1, 2):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    widths = {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return wb, ws


def update_actual(config: Any, trial_no: int, aktual_x_mm: float, aktual_y_mm: float) -> str:
    """Isi kolom Aktual X/Y (D/E) untuk baris `trial_no` yang sudah ada.

    Dipanggil segera setelah append_row(), supaya 1 baris langsung lengkap
    (Estimasi + Aktual + Error) tanpa perlu buka file XLSX secara manual.
    """
    from openpyxl import load_workbook

    path = _log_path(config)
    wb = load_workbook(path)
    ws = wb["Akurasi Titik"] if "Akurasi Titik" in wb.sheetnames else wb.active

    r = FIRST_DATA_ROW + int(trial_no) - 1
    ws.cell(row=r, column=4, value=round(float(aktual_x_mm), 2))
    ws.cell(row=r, column=5, value=round(float(aktual_y_mm), 2))

    wb.save(path)
    return path


def append_row(config: Any, estimasi_x_mm: float, estimasi_y_mm: float) -> tuple[str, int]:
    """Tambah 1 baris (Percobaan, Estimasi X, Estimasi Y) + rumus Error.

    Return (path, nomor_percobaan).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    path = _log_path(config)

    if not os.path.exists(path):
        wb, ws = _new_workbook(path)
    else:
        wb = load_workbook(path)
        ws = wb["Akurasi Titik"] if "Akurasi Titik" in wb.sheetnames else wb.active

    # Cari baris data kosong pertama (kolom A kosong) mulai FIRST_DATA_ROW.
    r = FIRST_DATA_ROW
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1

    trial_no = r - FIRST_DATA_ROW + 1
    font_name = "Times New Roman"
    body_font = Font(name=font_name, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=r, column=1, value=trial_no)
    ws.cell(row=r, column=2, value=round(float(estimasi_x_mm), 2))
    ws.cell(row=r, column=3, value=round(float(estimasi_y_mm), 2))
    # Kolom D/E (Aktual) sengaja dikosongkan untuk diisi manual.
    ws.cell(
        row=r,
        column=6,
        value=f'=IF(AND(B{r}<>"",C{r}<>"",D{r}<>"",E{r}<>""),SQRT((D{r}-B{r})^2+(E{r}-C{r})^2),"")',
    )
    for col in range(1, 7):
        cell = ws.cell(row=r, column=col)
        cell.font = body_font
        cell.alignment = center
        cell.border = border

    wb.save(path)
    return path, trial_no
